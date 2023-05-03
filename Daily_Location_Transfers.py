import os
import csv
from openpyxl import load_workbook
import pyautogui
#import cv2
import time
import requests
import json
import getpass





#*********************************
#IMPORT UNIT/LOCATION FROM VENDOR2
#*********************************


#this allows time to move the shell if it's blocking the Quick Menu on the Vendor2 window
time.sleep(5)


#assign the filepath to the current user's Documents folder.  Vendor2
#downloads the Activity Log to this folder by default.  Use the 
#jpeg copy of Vendor2's Quick Menu button to start the Vendor2 import process

filepath = os.path.expanduser('~\Documents\Activity Log Dashboard.xlsx')
username = getpass.getuser()
savepathstr = 'C:\\Users\\'+username+'\\Documents\\Activity Log Dashboard.xlsx'
quickmenuloc = pyautogui.locateCenterOnScreen('Vendor2_Quick_Menu.png')


print('Quick Menu button location: ', quickmenuloc, '\n')


#if the system can't find the Quick Menu button, the 
#user receives a warning message and the program stops

if not quickmenuloc:
    print("!!!PLEASE REPLACE THE SCREENSHOT OF Vendor2'S QUICK MENU BUTTON!!!")
    time.sleep(15)
    quit()


#delete any copies of the sheet in the current user's Download directory

try:
    os.remove(filepath)
    print('Old file deleted\n')
except OSError:
    print('No prior activity log detected\n')
    pass


print('Downloading new activity log from Vendor2\n')


#navigate an open Vendor2 session, download a new copy of the Activity Log

pyautogui.click(quickmenuloc)
time.sleep(2)
pyautogui.press('enter')
time.sleep(15)
pyautogui.press('tab',presses=5)
pyautogui.write('ng',interval=0.25)
pyautogui.press('enter')
time.sleep(7)
pyautogui.press('tab',presses=1)
pyautogui.press('enter')
time.sleep(35)
pyautogui.write(savepathstr, interval=0.1)
pyautogui.press('enter')
time.sleep(20)

print('New activity log downloaded\n')
print('Importing Vendor2 locations\n')


#navigate to the current user's Document directory to find the workbook

wb = load_workbook(filepath)
sheet = wb.active


#create a list of headers from row 4 of the active sheet
#this will be used to dynamically locate the columns we need

headers = []
for cell in sheet[4]:
    headers.append(cell.value)


#locate the unit number and location columns

units_col = headers.index("UN")
locations_col = headers.index("Location / Department")

Vendor2_locs = {}

print('Parsing Vendor2 locations\n')


#populate the Vendor2_locs dict with the unit numbers and locations

for r in sheet.iter_rows(min_row=5, max_row=1500, values_only=True):
    unit_num = r[units_col]
    loc_name = r[locations_col]
    try:
        Vendor2_locs.update({unit_num:loc_name.lower().rstrip()})
    except:
        if unit_num:
            Vendor2_locs.update({unit_num:'no loc name in Vendor2'})
        else:
            break


#there are a few naming conventions in our system to work around:

for k,v in Vendor2_locs.items():
    if v == 'crf':
        Vendor2_locs[k] = 'oklahoma city'
    elif v == 'okc':
        Vendor2_locs[k] = 'oklahoma city'


print('Vendor2 locations ready\n')

#print(Vendor2_locs)





#*********************************
#IMPORT UNIT/LOCATION FROM VENDOR1
#*********************************


print('Importing Vendor1 locations\n')


url = "https://api.Vendor1.com/industry/assets"

headers = {
    "Accept": "application/json",
    "Authorization": "Bearer Vendor1_api_insertbearertokenhere"
}

response = requests.request("GET", url, headers=headers)

#json result:

#{
#  "data": [
#    {
#      "parentAsset": {
#        "id": "cda345a1-0eac",
#        "name": "Customer Holdings, LLC"
#      },
#      "name": "Customer Holdings, LLC Southeast",
#      "id": "36789d1e-46fg",
#      "isRunning": false,
#      "customMetadata": {
#        "Dashboard Type": "Dashboard Test"
#      }
#    },  
#
#...
#
#],
#  "pagination": {
#    "endCursor": "4ed8c5a7-cd3e",
#    "hasNextPage": true
#  }
#}  etc....


#create a list of dictionaries from the JSON response
assets = json.loads(response.text)

#create a dictionary for the individual unit asset numbers 
Vendor1_locs = {}


#iterate over the JSON response to pull out unit asset numbers
#and their parent asset. Try/except is necessary 
#for the "Customer" asset, which doesn't have a parent asset

def update_loc_list():
    for asset in assets['data']:
        if 'G10' in asset['name']:
            unit = str(asset['name'])
            try:
                parent_asset = str(asset['parentAsset']['name'])
            except:
                parent_asset = 'No Parent Asset'
            Vendor1_locs.update({unit:parent_asset.lower()})

update_loc_list()


#the API only allows ~500 results in a page with responses.
#This section deals with pagination in the report.  It's not
#DRY though; the 'response' variable has to be modified.
#Can this section be improved?

while assets['pagination']['hasNextPage']:
    querystring = {'after':str(assets['pagination']['endCursor'])}
    response = requests.request("GET", url, headers=headers, params=querystring)
    assets = json.loads(response.text)

    update_loc_list()

print('Parsing Vendor1 locations\n')


#this section removes parts from the Vendor1 location names
#that aren't present in the Vendor2 location names.  It'll
#help reduce false negatives when we compare the two
#Additionally, it will add items to the Vendor2 dict if they're 
#in Vendor1 but not Vendor2 and avoid errors in the next phase.

for k,v in Vendor1_locs.items():
    if k not in Vendor2_locs:
        Vendor2_locs.update({k:"Unit not in Vendor2"})
    elif v[-5:] == ' site':
        Vendor1_locs[k] = v[:-5]
    elif v[-11:] == ' shop units':
        Vendor1_locs[k] = v[:-11]
    

#this section creates a copy of the 'Vendor1_locs' dict,
#then pops out the 'G10... Customer' units. They aren't 
#necessary when comparing locations in Vendor1 vs Vendor2

for unit_num in Vendor1_locs.copy():    
    if "Customer" in unit_num:
        Vendor1_locs.pop(unit_num)


print('Vendor1 locations ready\n')

#print(Vendor1_locs)





#**********************************************
#CONTRAST THE TWO DICTS, EXTRACT THE MISMATCHES
#**********************************************


print('Contrasting Vendor2 and Vendor1 locations\n')


#This list has a tuple that will be the header row
#Each unit/location pair that doesn't match will
#add another tuple to the list

c = [('Unit','Old Location','New Location')]


#The units(keys) are in the same order in each dict. If the location
#(values) don't match, the unit, old location, and new 
#location are made into a tuple and added to the list

diffkeys = [k for k in Vendor1_locs if Vendor1_locs[k] != Vendor2_locs[k]]
for k in diffkeys:
    c.append((k, Vendor1_locs[k], Vendor2_locs[k]))
           

#create (or overwrite) Sample.csv in the same folder as the script,
#and make a row for each tuple in the list.  Each value in 
#the tuple gets put into a separate column

with open('Unit_Transfers.csv','w', newline='') as output:
    writer = csv.writer(output)
    for t in c:
        writer.writerow([t[0],t[1],t[2]])

print('Program completed')